"""
user/lark_bot/ —— 飞书/Lark Bot 用户交互实现。

基于 lark-oapi SDK 实现飞书机器人接口。
使用 WebSocket 长连接模式，无需公网 IP。
"""

from __future__ import annotations

import asyncio
import json
import threading
import time
from typing import TYPE_CHECKING

import lark_oapi as lark
from lark_oapi.api.im.v1 import (
    CreateMessageRequest,
    CreateMessageRequestBody,
)

from logger import log, new_log
from user.user import BaseUser
from config import get_config

if TYPE_CHECKING:
    from host.momoka import Momoka


class LarkBotUser(BaseUser):
    """基于飞书/Lark Bot 的用户交互实现。

    支持:
    - 一对一聊天
    - 群聊 @提及 回复
    - 命令交互
    
    使用 WebSocket 长连接模式，无需公网 IP 或 verification_token。
    """
    
    interface_type = 'lark'

    def __init__(
        self,
        app_id: str,
        app_secret: str,
    ):
        """
        Args:
            app_id: 飞书应用 App ID
            app_secret: 飞书应用 App Secret
        """
        super().__init__()
        self._agent: Momoka | None = None
        self._start_time = 0.0
        self._app_id = app_id
        self._app_secret = app_secret

        # 消息队列: open_chat_id -> list[message]
        self._message_queues: dict[str, list[str]] = {}
        self._queues_lock = threading.Lock()

        # 飞书 Client
        self._client: lark.Client | None = None

        # 当前活跃的聊天 ID
        self._active_chat_id: str | None = None
        
        # 是否已发送过欢迎消息(每个聊天独立追踪)
        self._welcome_sent_chats: set[str] = set()

    def set_agent(self, agent: Momoka) -> None:
        """设置关联的 Agent 实例。"""
        self._agent = agent

    def _get_queue(self, chat_id: str) -> list[str]:
        """获取或创建消息队列。"""
        with self._queues_lock:
            if chat_id not in self._message_queues:
                self._message_queues[chat_id] = []
            return self._message_queues[chat_id]

    def _put_message(self, chat_id: str, message: str) -> None:
        """将消息放入队列。"""
        queue = self._get_queue(chat_id)
        with self._queues_lock:
            queue.append(message)

    def _has_message(self, chat_id: str) -> bool:
        """检查是否有新消息。"""
        with self._queues_lock:
            return chat_id in self._message_queues and len(self._message_queues[chat_id]) > 0

    def _get_message(self, chat_id: str) -> str | None:
        """从队列获取消息。"""
        with self._queues_lock:
            queue = self._message_queues.get(chat_id, [])
            if queue:
                return queue.pop(0)
            return None

    def _send_welcome_and_system(self, chat_id: str) -> None:
        """发送欢迎消息到用户和 system 提示给 Agent。

        Args:
            chat_id: 聊天 ID
        """
        if chat_id in self._welcome_sent_chats:
            return
        
        self._welcome_sent_chats.add(chat_id)
        
        # 发送欢迎消息到用户
        welcome_msg = (
            "Successfully connected to Lark\n"
            "Welcome back! This is Momoka~\n"
            "Developed by Mikoris | For more help, type /help"
        )
        self._send_lark_message(chat_id, welcome_msg)
        log(f'lark | welcome message sent to chat {chat_id}')
        
        # 发送 system 消息给 Agent(添加到对话历史,不触发响应)
        if self._agent is not None:
            system_msg = "<You are communicating with user via Lark>"
            self._agent._model._ctx.history.append({
                'role': 'system',
                'content': system_msg
            })
            self._agent._model._ctx._meta.append({})
            log(f'lark | system message added to agent history')

    def _get_client(self) -> lark.Client:
        """获取飞书 Client。"""
        if self._client is None:
            self._client = (
                lark.Client.builder()
                .app_id(self._app_id)
                .app_secret(self._app_secret)
                .build()
            )
        return self._client

    def _send_lark_message(self, chat_id: str, text: str) -> None:
        """发送消息到飞书(同步版本)。

        Args:
            chat_id: 聊天 ID (open_chat_id 或 open_id)
            text: 消息文本
        """
        try:
            client = self._get_client()

            # 构建消息体
            msg_content = json.dumps({"text": text})

            # 创建发送请求
            request = (
                CreateMessageRequest.builder()
                .receive_id_type('chat_id')
                .request_body(
                    CreateMessageRequestBody.builder()
                    .receive_id(chat_id)
                    .msg_type('text')
                    .content(msg_content)
                    .build()
                )
                .build()
            )

            # 发送消息
            response = client.im.v1.message.create(request)
            if not response.success():
                log(f'lark | send message failed: {response.code} {response.msg}')
            else:
                log(f'lark | message sent to {chat_id}')
        except Exception as e:
            log(f'lark | send error: {e}')

    def _handle_message_event(self, data: lark.P2ImMessageReceiveV1) -> None:
        """处理接收到的飞书消息。

        Args:
            data: 消息事件对象
        """
        try:
            event = data.event
            message = event.message

            # 获取消息内容
            msg_type = message.message_type
            chat_id = message.chat_id

            if msg_type == 'text':
                content = message.content
                # content 是 JSON 字符串
                try:
                    content_obj = json.loads(content)
                    text = content_obj.get('text', '')
                except:
                    text = content

                # 去除 @提及
                if text.startswith('@_user_'):
                    parts = text.split(' ', 1)
                    if len(parts) > 1:
                        text = parts[1].strip()

                log(f'lark | receive from {chat_id}: {text}')
                self._put_message(chat_id, text)
            elif msg_type == 'file':
                # 处理文件消息
                self._handle_file_message(message, chat_id)
            else:
                log(f'lark | unsupported message type: {msg_type}')
        except Exception as e:
            log(f'lark | parse message error: {e}')

    def _handle_file_message(self, message, chat_id: str) -> None:
        """处理接收到的文件消息。

        Args:
            message: 飞书消息对象
            chat_id: 聊天 ID
        """
        try:
            content = message.content
            message_id = message.message_id
            try:
                content_obj = json.loads(content)
                file_key = content_obj.get('file_key', '')
                file_name = content_obj.get('file_name', 'unknown_file')
            except:
                log(f'lark | failed to parse file message content')
                return

            if not file_key:
                log(f'lark | file_key is empty')
                return

            log(f'lark | receive file from {chat_id}: {file_name} (key: {file_key}, msg_id: {message_id})')

            # 下载文件（使用消息资源接口，因为文件是用户发送的）
            file_path = self._download_lark_file_from_message(message_id, file_key, file_name)
            if file_path:
                # 读取文件内容
                file_content = self._read_file_content(file_path, file_name)
                if file_content:
                    # 将文件内容作为消息放入队列
                    message_text = f"[User uploaded file: {file_name}]\n\n{file_content}"
                    self._put_message(chat_id, message_text)
                    log(f'lark | file content queued: {file_name}')
                else:
                    self._put_message(chat_id, f"[User uploaded file: {file_name}]\n[File downloaded to: {file_path} but could not be read as text]")
            else:
                log(f'lark | failed to download file: {file_name}')
                self._send_lark_message(chat_id, f"Failed to receive file: {file_name}")
        except Exception as e:
            log(f'lark | handle file message error: {e}')
            import traceback
            log(f'lark | traceback: {traceback.format_exc()}')

    def _download_lark_file_from_message(self, message_id: str, file_key: str, file_name: str) -> str | None:
        """从飞书消息中下载资源文件（用于下载用户发送的文件）。

        注意：下载用户发送的文件需要使用 /messages/{message_id}/resources/{file_key} 接口，
        而不是 /files/{file_key} 接口。后者只能下载机器人自己上传的文件。

        Args:
            message_id: 飞书消息 ID
            file_key: 飞书文件 key
            file_name: 文件名

        Returns:
            下载后的本地文件路径，失败返回 None
        """
        try:
            import os
            import tempfile
            import uuid
            import requests

            # 获取 tenant_access_token
            token = self._get_tenant_access_token()
            if not token:
                log(f'lark | failed to get tenant_access_token')
                return None

            # 构建下载 URL（使用消息资源接口）
            url = f'https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{file_key}?type=file'

            # 发送下载请求
            headers = {
                'Authorization': f'Bearer {token}'
            }
            response = requests.get(url, headers=headers, timeout=60)

            if response.status_code != 200:
                log(f'lark | download file failed: {response.status_code} {response.text}')
                return None

            # 保存文件到临时目录
            work_dir = get_config().get('work_dir', tempfile.gettempdir())
            temp_dir = os.path.join(work_dir, 'temp')
            os.makedirs(temp_dir, exist_ok=True)

            # 生成唯一文件名避免冲突
            unique_name = f"{uuid.uuid4().hex[:8]}_{file_name}"
            file_path = os.path.join(temp_dir, unique_name)

            # 写入文件
            with open(file_path, 'wb') as f:
                f.write(response.content)

            log(f'lark | file downloaded to: {file_path}')
            return file_path

        except Exception as e:
            log(f'lark | download file error: {e}')
            import traceback
            log(f'lark | traceback: {traceback.format_exc()}')
            return None

    def _get_tenant_access_token(self) -> str | None:
        """获取飞书 tenant_access_token。

        Returns:
            token 字符串，失败返回 None
        """
        try:
            import requests
            import json

            url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
            headers = {
                'Content-Type': 'application/json; charset=utf-8'
            }
            payload = {
                'app_id': self._app_id,
                'app_secret': self._app_secret
            }

            response = requests.post(url, headers=headers, data=json.dumps(payload), timeout=30)
            result = response.json()

            if result.get('code') != 0:
                log(f'lark | get tenant_access_token failed: {result}')
                return None

            return result.get('tenant_access_token')

        except Exception as e:
            log(f'lark | get tenant_access_token error: {e}')
            return None

    def _read_file_content(self, file_path: str, file_name: str) -> str | None:
        """读取文件内容为文本。

        Args:
            file_path: 文件路径
            file_name: 文件名

        Returns:
            文件文本内容，如果不是文本文件返回 None
        """
        import os

        # 检查文件大小（限制 10MB）
        max_size = 10 * 1024 * 1024  # 10MB
        file_size = os.path.getsize(file_path)
        if file_size > max_size:
            return f"[File too large: {file_size / 1024 / 1024:.2f}MB, max 10MB]"

        # 根据扩展名判断文件类型
        ext = os.path.splitext(file_name)[1].lower()

        # 文本文件扩展名
        text_extensions = {
            '.txt', '.md', '.py', '.js', '.ts', '.html', '.css', '.json',
            '.xml', '.yaml', '.yml', '.csv', '.log', '.ini', '.conf',
            '.sh', '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd',
            '.c', '.cpp', '.h', '.hpp', '.java', '.go', '.rs', '.swift',
            '.kt', '.scala', '.rb', '.php', '.pl', '.lua', '.r', '.m',
            '.sql', '.dockerfile', '.makefile', '.cmake', '.gradle',
            '.vue', '.jsx', '.tsx', '.svelte', '.less', '.scss', '.sass',
        }

        # Office 文档扩展名（需要特殊处理）
        office_extensions = {
            '.docx', '.xlsx', '.pptx',
        }

        if ext in office_extensions:
            return self._read_office_file(file_path, ext)

        # 尝试作为文本文件读取
        if ext in text_extensions or ext not in {'.pdf', '.doc', '.xls', '.ppt', '.zip', '.rar', '.7z', '.tar', '.gz', '.exe', '.dll', '.so', '.dylib'}:
            return self._try_read_text_file(file_path)

        return None

    def _try_read_text_file(self, file_path: str) -> str | None:
        """尝试以文本方式读取文件。

        Args:
            file_path: 文件路径

        Returns:
            文件内容，如果读取失败返回 None
        """
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1', 'cp1252']

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                return content
            except UnicodeDecodeError:
                continue
            except Exception as e:
                log(f'lark | read file error with {encoding}: {e}')
                continue

        return None

    def _read_office_file(self, file_path: str, ext: str) -> str | None:
        """读取 Office 文档内容。

        Args:
            file_path: 文件路径
            ext: 文件扩展名

        Returns:
            文档文本内容
        """
        try:
            if ext == '.docx':
                from docx import Document
                doc = Document(file_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return '\n'.join(paragraphs)
            elif ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                lines = []
                for sheet in wb.worksheets:
                    lines.append(f'--- Sheet: {sheet.title} ---')
                    for row in sheet.iter_rows(values_only=True):
                        row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                        if row_text.strip():
                            lines.append(row_text)
                return '\n'.join(lines)
            elif ext == '.pptx':
                # python-pptx 需要单独安装，这里简单返回提示
                return "[PPTX file - content extraction not supported]"
        except Exception as e:
            log(f'lark | read office file error: {e}')
            return f"[Failed to read {ext} file: {e}]"

        return None

    def run(self) -> None:
        """启动飞书 Bot 会话循环。"""
        if self._agent is None:
            raise RuntimeError("Agent not set. Call set_agent() before run().")

        new_log()
        log('lark_bot_start')

        self._start_time = time.time()
        self.session.reset()

        # 定义消息接收回调
        def on_message(data: lark.P2ImMessageReceiveV1) -> None:
            self._handle_message_event(data)

        # 构建事件分发器 (WebSocket 长连接模式)
        event_handler = (
            lark.EventDispatcherHandler.builder("", "")
            .register_p2_im_message_receive_v1(on_message)
            .build()
        )

        # 创建 WebSocket 长连接客户端 (新版本 SDK 直接初始化)
        ws_client = lark.ws.Client(
            app_id=self._app_id,
            app_secret=self._app_secret,
            event_handler=event_handler,
            log_level=lark.LogLevel.DEBUG,
        )

        def run_ws():
            """在独立线程中运行 WebSocket 连接。"""
            log('lark | WebSocket connecting...')
            ws_client.start()

        # 启动 WebSocket 连接
        ws_thread = threading.Thread(target=run_ws, daemon=True)
        ws_thread.start()

        # 等待连接建立
        time.sleep(2)

        # 主消息循环
        self._run_main_loop()

    def _run_main_loop(self) -> None:
        """主消息循环: 轮询消息队列,调用 Agent,返回结果。"""
        # 创建斜杠命令处理器
        from user.lark_bot.interactions import LarkInteractionAdapter
        adapter = LarkInteractionAdapter(self)
        slash_handler = self._create_slash_handler(adapter)

        # 模型选择状态
        waiting_for_model_choice = False
        pending_models: list[str] = []

        # 清空确认状态
        waiting_for_clear_confirmation = False

        while True:
            # 检查所有队列
            active_chat_id = None
            for chat_id in self._message_queues:
                if self._has_message(chat_id):
                    active_chat_id = chat_id
                    break

            if active_chat_id is None:
                time.sleep(0.1)
                continue

            user_message = self._get_message(active_chat_id)
            if user_message is None:
                continue

            # 设置当前活跃的聊天 ID
            self._active_chat_id = active_chat_id

            # 首条消息:发送欢迎消息和 system 提示
            self._send_welcome_and_system(active_chat_id)

            # 如果正在等待模型选择，处理用户的数字回复
            if waiting_for_model_choice:
                waiting_for_model_choice = False
                if self._handle_model_choice(user_message, pending_models):
                    continue
                # 如果选择无效，继续作为普通消息处理

            # 如果正在等待清空确认，处理用户的回复
            if waiting_for_clear_confirmation:
                waiting_for_clear_confirmation = False
                if self._handle_clear_confirmation(user_message):
                    continue

            if self._agent is None:
                self._send_lark_message(active_chat_id, "Error: Agent not initialized")
                continue

            # 检查是否为斜杠命令
            if user_message.strip().startswith('/'):
                # 如果是 /model 命令，需要特殊处理
                if user_message.strip() == '/model':
                    models = self._fetch_models_for_model()
                    if models:
                        pending_models = models
                        waiting_for_model_choice = True
                        # 发送模型列表
                        cfg = get_config()
                        current = cfg.get('model', '')
                        model_list = '\n'.join([
                            f"  {i+1}. {m}{' [current]' if m == current else ''}"
                            for i, m in enumerate(models)
                        ])
                        self._send_lark_message(
                            active_chat_id,
                            f"Available models:\n{model_list}\n\n"
                            f"Please reply with the model number (1-{len(models)}) or leave blank to cancel."
                        )
                        continue
                    else:
                        self._send_lark_message(active_chat_id, 'No models available or failed to fetch.')
                        continue

                handled, skill_name = slash_handler.handle(user_message)
                if handled:
                    # /end 命令特殊处理
                    if skill_name == '__end__':
                        self._handle_end_command()
                        break
                    # /init 命令
                    if skill_name == '__init__':
                        continue
                    # /clear 命令（等待用户确认）
                    if skill_name == '__clear_ask__':
                        waiting_for_clear_confirmation = True
                        continue
                    # Skill 加载
                    if skill_name is not None:
                        log(f'lark | skill trigger: {skill_name}')
                        load_result = self._agent.load_skill(skill_name)
                        if load_result.success:
                            self._send_lark_message(active_chat_id, f'Load Skill: {skill_name}')
                            log(f'lark (skill inject): {skill_name}')
                        else:
                            self._send_lark_message(active_chat_id, f'Non-existent command or skill: {skill_name}')
                        continue
                    # 其他命令已处理
                    continue

            # 修复历史
            repaired = self._agent.repair_history()
            if repaired:
                log(f'lark | repair_history: filled in {repaired} orphaned tool_calls')

            # 调用 Agent
            try:
                result = self._agent.send(
                    user_message,
                    file_contents=self.session.file_contents
                )

                self.session.update(result)

                if result.is_finish:
                    self._agent.finish_task()
                    self.on_task_finish()

            except Exception as e:
                log(f'lark | agent error: {e}')
                self._send_lark_message(active_chat_id, f"Error: \n{str(e)}")

    def get_input(self) -> str:
        """等待并获取用户输入(阻塞式)。"""
        while True:
            for chat_id in self._message_queues:
                msg = self._get_message(chat_id)
                if msg:
                    return msg
            time.sleep(0.1)

    def send_output(self, message: str, role: str = 'BOT') -> None:
        """向用户输出消息。"""
        log(f'lark ({role}) | {message}')

        # 如果有活跃的聊天 ID，发送消息到飞书
        if self._active_chat_id:
            if role == 'BOT':
                self._send_lark_message(self._active_chat_id, message)
            else:
                self._send_lark_message(self._active_chat_id, f"[{role}]\n{message}")

    def send_error(self, message: str) -> None:
        """输出错误消息。"""
        log(f'lark ERROR | {message}')

    def send_file(self, file_path: str, caption: str = '') -> None:
        """向用户发送一个文件（飞书支持文件消息）。

        Args:
            file_path: 文件的绝对路径
            caption: 可选的伴随消息
        """
        import os

        log(f'lark | send_file: {file_path} | caption: {caption}')

        # 验证文件是否存在
        if not os.path.exists(file_path):
            log(f'lark | send_file error: file not found: {file_path}')
            self.send_error(f'File not found: {file_path}')
            return

        # 如果有活跃的聊天 ID，发送文件到飞书
        if self._active_chat_id:
            self._send_lark_file(self._active_chat_id, file_path, caption)

    def _send_lark_file(self, chat_id: str, file_path: str, caption: str = '') -> None:
        """发送文件到飞书（先上传获取 file_key，再发送文件消息）。

        Args:
            chat_id: 聊天 ID
            file_path: 文件路径
            caption: 可选的伴随消息
        """
        import os

        try:
            client = self._get_client()
            file_name = os.path.basename(file_path)

            log(f'lark | uploading file: {file_name}')

            # 判断文件类型
            file_ext = os.path.splitext(file_name)[1].lower()
            image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
            is_image = file_ext in image_extensions

            if is_image:
                # 图片文件：使用 image API
                from lark_oapi.api.im.v1 import (
                    CreateImageRequest,
                    CreateImageRequestBody,
                )

                # 图片 API 需要重新打开文件
                with open(file_path, 'rb') as img_file:
                    upload_request = (
                        CreateImageRequest.builder()
                        .request_body(
                            CreateImageRequestBody.builder()
                            .image_type('message')
                            .image(img_file)
                            .build()
                        )
                        .build()
                    )

                    upload_response = client.im.v1.image.create(upload_request)

                if not upload_response.success():
                    log(f'lark | image upload failed: {upload_response.code} {upload_response.msg}')
                    self.send_error(f'Failed to upload image: {upload_response.msg}')
                    return

                image_key = upload_response.data.image_key
                log(f'lark | image uploaded, image_key: {image_key}')

                # 发送图片消息
                msg_content = json.dumps({"image_key": image_key})
                msg_type = 'image'
            else:
                # 普通文件：使用 file API
                from lark_oapi.api.im.v1 import (
                    CreateFileRequest,
                    CreateFileRequestBody,
                )

                # 使用 'stream' 作为通用文件类型（飞书官方推荐）
                # 参考: https://open.feishu.cn/document/server-docs/im-v1/file/create
                file_type = 'stream'

                # 重新打开文件，传入文件对象而不是 bytes
                with open(file_path, 'rb') as file_stream:
                    upload_request = (
                        CreateFileRequest.builder()
                        .request_body(
                            CreateFileRequestBody.builder()
                            .file_type(file_type)
                            .file_name(file_name)
                            .file(file_stream)
                            .build()
                        )
                        .build()
                    )

                    upload_response = client.im.v1.file.create(upload_request)

                if not upload_response.success():
                    log(f'lark | file upload failed: {upload_response.code} {upload_response.msg}')
                    self.send_error(f'Failed to upload file: {upload_response.msg}')
                    return

                file_key = upload_response.data.file_key
                log(f'lark | file uploaded, file_key: {file_key}')

                # 发送文件消息
                msg_content = json.dumps({"file_key": file_key})
                msg_type = 'file'

            # 发送消息
            from lark_oapi.api.im.v1 import (
                CreateMessageRequest as LarkCreateMessageRequest,
                CreateMessageRequestBody as LarkCreateMessageRequestBody,
            )

            message_request = (
                LarkCreateMessageRequest.builder()
                .receive_id_type('chat_id')
                .request_body(
                    LarkCreateMessageRequestBody.builder()
                    .receive_id(chat_id)
                    .msg_type(msg_type)
                    .content(msg_content)
                    .build()
                )
                .build()
            )

            message_response = client.im.v1.message.create(message_request)

            if not message_response.success():
                log(f'lark | send file message failed: {message_response.code} {message_response.msg}')
                self.send_error(f'Failed to send file message: {message_response.msg}')
                return

            log(f'lark | file sent: {file_name}')

            # 如果有 caption，再发送一条文本消息
            if caption:
                self._send_lark_message(chat_id, caption)

        except Exception as e:
            log(f'lark | send file error: {e}')
            import traceback
            log(f'lark | traceback: {traceback.format_exc()}')
            self.send_error(f'Failed to send file: {e}')

    def on_task_finish(self) -> None:
        """任务完成回调。"""
        log('lark | task finished')

    def on_session_end(self, input_tokens: int, output_tokens: int,
                       round_count: int, elapsed: float) -> None:
        """会话结束回调。"""
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)
        time_str = f'{mins}min {secs}s' if mins else f'{secs}s'
        log(f'lark | session end ({time_str} | Input: {input_tokens} tokens | Output: {output_tokens} tokens | {round_count}R)')
    
    def _create_slash_handler(self, adapter) -> 'SlashCommandHandler':
        """创建斜杠命令处理器。"""
        from user.commands import SlashCommandHandler
        return SlashCommandHandler(adapter.create_callbacks())
    
    def _handle_end_command(self) -> None:
        """处理 /end 命令。"""
        elapsed = time.time() - self._start_time
        self.on_session_end(
            self.session.input_tokens,
            self.session.output_tokens,
            self.session.round_count,
            elapsed
        )

    def _fetch_models_for_model(self) -> list[str]:
        """为 /model 命令获取模型列表。"""
        try:
            from openai import OpenAI
            cfg = get_config()
            client = OpenAI(api_key=cfg['api_key'], base_url=cfg['base_url'])
            models = client.models.list()
            return sorted(m.id for m in models.data)
        except Exception as e:
            log(f'lark | failed to fetch models: {e}')
            return []

    def _handle_model_choice(self, user_message: str, models: list[str]) -> bool:
        """处理用户的模型选择输入。

        Args:
            user_message: 用户回复的消息
            models: 模型列表

        Returns:
            True 表示已处理，False 表示输入无效
        """
        user_input = user_message.strip()

        # 空输入视为取消
        if not user_input:
            self._send_lark_message(self._active_chat_id, 'Cancelled.')
            return True

        # 尝试解析数字
        try:
            choice = int(user_input)
        except ValueError:
            # 不是数字，视为无效，返回 False 让主循环继续作为普通消息处理
            return False

        # 验证范围
        if choice < 1 or choice > len(models):
            self._send_lark_message(
                self._active_chat_id,
                f'Invalid selection. Please enter a number between 1 and {len(models)}.'
            )
            return True

        selected_model = models[choice - 1]
        cfg = get_config()
        current = cfg.get('model', '')

        # 如果选择的模型和当前相同，无需更新
        if selected_model == current:
            self._send_lark_message(
                self._active_chat_id,
                f'Model unchanged: {selected_model}'
            )
            return True

        # 更新配置
        try:
            import json
            from config import CONFIG_FILE
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config_data = json.load(f)

            config_data['model'] = selected_model

            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=2)

            self._send_lark_message(
                self._active_chat_id,
                f'Model updated: {current!r} → {selected_model!r}'
            )
            log(f'lark | model changed: {current} → {selected_model}')
            return True

        except Exception as e:
            self._send_lark_message(
                self._active_chat_id,
                f'Failed to save model: {e}'
            )
            return True

    def _handle_clear_confirmation(self, user_message: str) -> bool:
        """处理用户对清空操作的确认回复。

        Args:
            user_message: 用户回复的消息

        Returns:
            True 表示已处理
        """
        user_input = user_message.strip().lower()

        # 检查是否确认
        if user_input in ('y', 'yes'):
            try:
                # 清空模型层的对话历史
                if self._agent:
                    self._agent.clear_context()

                    # 重置会话状态（token 统计等）
                    self.session.reset()

                    # 触发回调
                    self.on_clear_context()

                self._send_lark_message(self._active_chat_id, 'Context cleared.')
                return True

            except Exception as e:
                self._send_lark_message(self._active_chat_id, f'Failed to clear context: {e}')
                return True
        else:
            # 用户取消
            self._send_lark_message(self._active_chat_id, 'Cancelled.')
            return True
