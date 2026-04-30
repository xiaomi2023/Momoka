"""
user/qq_bot/qq_bot.py —— QQ Bot 用户交互实现。

基于 QQ开放平台 WebSocket 网关实现 QQ 机器人接口。
支持群聊、私聊消息收发，自动处理 Access Token 鉴权。
"""

from __future__ import annotations

import asyncio
import json
import os
import threading
import time
from typing import TYPE_CHECKING

import websockets
import aiohttp

from logger import log, new_log
from user.user import BaseUser
from config import get_config

if TYPE_CHECKING:
    from host.momoka import Momoka


class QQBotUser(BaseUser):
    """基于 QQ开放平台 WebSocket 网关的 QQ Bot 用户交互实现。

    支持:
    - 群聊 @提及 回复
    - 私聊对话
    - 消息队列机制
    - Access Token 自动刷新
    """

    interface_type = 'qq'

    def __init__(self, app_id: str, app_secret: str, sandbox: bool = False):
        """
        Args:
            app_id: QQ开放平台应用 ID
            app_secret: QQ开放平台应用密钥
            sandbox: 是否使用沙箱环境
        """
        super().__init__()
        self._agent: Momoka | None = None
        self._start_time = 0.0
        self._app_id = app_id
        self._app_secret = app_secret
        self._sandbox = sandbox

        # Access Token 管理
        self._access_token: str | None = None
        self._token_expires_at: float = 0.0

        # WebSocket 连接
        self._ws_url: str = ''
        self._session_id: str = ''
        self._seq: int = 0
        self._session_start_limit: dict = {}
        self._ws: websockets.WebSocketClientProtocol | None = None

        # 消息队列: chat_id -> asyncio.Queue
        self._message_queues: dict[str, asyncio.Queue] = {}

        # 新消息事件（用于替代轮询）
        self._message_event: asyncio.Event | None = None

        # 当前活跃聊天 ID
        self._active_chat_id: str | None = None

        # 事件循环和线程
        self._event_loop: asyncio.AbstractEventLoop | None = None
        self._ws_thread: threading.Thread | None = None

        # 是否已发送过欢迎消息
        self._welcome_sent_chats: set[str] = set()

        # Bot 运行标志
        self._running = False

        # WebSocket 重连控制
        self._reconnect_attempts: int = 0
        self._max_reconnect_attempts: int = 5

    def set_agent(self, agent: Momoka) -> None:
        """设置关联的 Agent 实例。"""
        self._agent = agent

    def _get_api_base(self) -> str:
        """获取 API 基础 URL。"""
        if self._sandbox:
            return 'https://sandbox.api.sgroup.qq.com'
        return 'https://api.sgroup.qq.com'

    def _get_ws_base(self) -> str:
        """获取 WebSocket 基础 URL。"""
        if self._sandbox:
            return 'wss://sandbox.api.sgroup.qq.com'
        return 'wss://api.sgroup.qq.com'

    async def _get_access_token(self) -> str:
        """获取或刷新 Access Token。

        Returns:
            有效的 Access Token
        """
        # 检查 token 是否还有效（提前 60 秒刷新）
        if self._access_token and time.time() < self._token_expires_at - 60:
            return self._access_token

        log('qq | refreshing access token')

        url = 'https://bots.qq.com/app/getAppAccessToken'
        payload = {
            'appId': self._app_id,
            'clientSecret': self._app_secret
        }

        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=payload) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    raise Exception(f'Failed to get access token: {resp.status} {error_text}')

                data = await resp.json()
                self._access_token = data['access_token']
                expires_in = int(data.get('expires_in', 7200))
                self._token_expires_at = time.time() + expires_in

                log(f'qq | access token refreshed, expires in {expires_in}s')
                return self._access_token

    def _get_auth_header(self) -> dict:
        """获取鉴权请求头。"""
        token = self._access_token or 'invalid'
        return {
            'Authorization': f'QQBot {token}',
            'Content-Type': 'application/json'
        }

    async def _get_gateway_url(self) -> str:
        """从开放平台获取 WebSocket 网关 URL。

        Returns:
            WebSocket 连接 URL
        """
        token = await self._get_access_token()

        url = f'{self._get_api_base()}/gateway/bot'
        headers = self._get_auth_header()

        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    raise Exception(f'Failed to get gateway: {resp.status} {error_text}')

                data = await resp.json()
                ws_url = data.get('url', f'{self._get_ws_base()}/websocket/')
                self._session_start_limit = data.get('session_start_limit', {})

                log(f'qq | gateway url: {ws_url}, shards: {data.get("shards", 1)}')
                return ws_url

    async def _send_ws(self, payload: dict) -> None:
        """发送数据到 WebSocket。

        Args:
            payload: 要发送的 JSON 数据
        """
        if self._ws:
            await self._ws.send(json.dumps(payload))

    async def _identify(self) -> None:
        """发送 Identify 鉴权消息。"""
        token = await self._get_access_token()

        # 计算 intents: 订阅群消息(1<<25) + 公域消息(1<<30) + C2C消息(1<<25)
        # 0 = 基础事件, 1<<25 = 群/单聊, 1<<30 = 公域
        intents = 0 | (1 << 25) | (1 << 30)

        payload = {
            'op': 2,  # Identify
            'd': {
                'token': f'QQBot {token}',
                'intents': intents,
                'shard': [0, 1],
                'properties': {
                    '$os': 'windows',
                    '$browser': 'momoka',
                    '$device': 'momoka'
                }
            }
        }

        await self._send_ws(payload)
        log('qq | identify sent')

    async def _send_heartbeat(self) -> None:
        """发送心跳包。"""
        payload = {
            'op': 1,  # Heartbeat
            'd': self._seq
        }
        await self._send_ws(payload)

    def _get_queue(self, chat_id: str) -> asyncio.Queue:
        """获取或创建指定聊天的消息队列。

        Args:
            chat_id: 聊天 ID (群号/用户openid)

        Returns:
            消息队列
        """
        if chat_id not in self._message_queues:
            self._message_queues[chat_id] = asyncio.Queue()
        return self._message_queues[chat_id]

    def _notify_message_received(self) -> None:
        """通知主循环有新消息到达。"""
        if self._message_event and self._event_loop and self._event_loop.is_running():
            # asyncio.Event.set() 是线程安全的，可以直接跨线程调用
            self._message_event.set()

    async def _handle_attachment(self, attachment: dict, chat_id: str, msg_id: str, author_id: str) -> None:
        """处理单个附件（用户发送的文件）。

        Args:
            attachment: 附件信息字典，包含 filename, content_type, size, url 等
            chat_id: 聊天 ID
            msg_id: 消息 ID
            author_id: 发送者 ID
        """
        try:
            filename = attachment.get('filename', 'unknown_file')
            content_type = attachment.get('content_type', 'unknown')
            size = attachment.get('size', 0)
            url = attachment.get('url', '')

            if not url:
                log(f'qq | attachment url is empty, skip download: {filename}')
                return

            log(f'qq | receive file from {chat_id}: {filename} (type: {content_type}, size: {size}, url: {url})')

            # 下载文件
            file_path = await self._download_file_from_url(url, filename)
            if file_path:
                # 读取文件内容
                file_content = self._read_file_content(file_path, filename)
                if file_content:
                    # 将文件内容作为消息放入队列
                    message_text = f"<{filename}>\n{file_content}\n</{filename}>"
                    queue = self._get_queue(chat_id)
                    await queue.put({
                        'chat_id': chat_id,
                        'msg_id': msg_id,
                        'content': message_text,
                        'author_id': author_id,
                        'type': 'file'
                    })
                    self._notify_message_received()
                    log(f'qq | file content queued: {filename}')
                else:
                    # 文件下载成功但无法读取为文本
                    queue = self._get_queue(chat_id)
                    await queue.put({
                        'chat_id': chat_id,
                        'msg_id': msg_id,
                        'content': f"<{filename}>\n[File downloaded to: {file_path} but could not be read as text]\n</{filename}>",
                        'author_id': author_id,
                        'type': 'file'
                    })
                    self._notify_message_received()
                    log(f'qq | file downloaded but not readable: {filename}')
            else:
                log(f'qq | failed to download file: {filename}')
                self._send_qq_message_sync(chat_id, f"Failed to receive file: {filename}")
        except Exception as e:
            log(f'qq | handle attachment error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    async def _download_file_from_url(self, url: str, file_name: str) -> str | None:
        """从 URL 下载文件到本地临时目录。

        Args:
            url: 文件下载链接
            file_name: 文件名

        Returns:
            下载后的本地文件路径，失败返回 None
        """
        try:
            import os
            import tempfile
            import uuid
            import aiohttp

            # 构建下载 URL（可能需要添加认证头）
            download_url = url

            # 发送下载请求
            headers = {
                'User-Agent': 'QQBot/1.0'
            }

            async with aiohttp.ClientSession() as session:
                async with session.get(download_url, headers=headers, timeout=aiohttp.ClientTimeout(total=120)) as response:
                    if response.status != 200:
                        log(f'qq | download file failed: {response.status}')
                        return None

                    # 保存文件到临时目录
                    work_dir = get_config().get('work_dir', tempfile.gettempdir())
                    temp_dir = os.path.join(work_dir, 'temp')
                    os.makedirs(temp_dir, exist_ok=True)

                    # 生成唯一文件名避免冲突
                    unique_name = f"{uuid.uuid4().hex[:8]}_{file_name}"
                    file_path = os.path.join(temp_dir, unique_name)

                    # 写入文件
                    file_data = await response.read()
                    with open(file_path, 'wb') as f:
                        f.write(file_data)

                    log(f'qq | file downloaded to: {file_path}')
                    return file_path

        except Exception as e:
            log(f'qq | download file error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')
            return None

    def _read_file_content(self, file_path: str, file_name: str) -> str | None:
        """读取文件内容为文本。

        Args:
            file_path: 文件路径
            file_name: 文件名

        Returns:
            文件文本内容，如果不是文本文件返回 None
        """
        import os

        # 检查文件大小（限制 10MB）
        max_size = 10 * 1024 * 1024  # 10MB
        try:
            file_size = os.path.getsize(file_path)
            if file_size > max_size:
                return f"[File too large: {file_size / 1024 / 1024:.2f}MB, max 10MB]"
        except:
            return None

        # 根据扩展名判断文件类型
        ext = os.path.splitext(file_name)[1].lower()

        # 文本文件扩展名
        text_extensions = {
            '.txt', '.md', '.py', '.js', '.ts', '.html', '.css', '.json',
            '.xml', '.yaml', '.yml', '.csv', '.log', '.ini', '.conf',
            '.sh', '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd',
            '.c', '.cpp', '.h', '.hpp', '.java', '.go', '.rs', '.swift',
            '.kt', '.scala', '.rb', '.php', '.pl', '.lua', '.r', '.m',
            '.sql', '.dockerfile', '.makefile', '.cmake', '.gradle',
            '.vue', '.jsx', '.tsx', '.svelte', '.less', '.scss', '.sass',
        }

        # Office 文档扩展名（需要特殊处理）
        office_extensions = {
            '.docx', '.xlsx',
        }

        if ext in office_extensions:
            return self._read_office_file(file_path, ext)

        # 尝试作为文本文件读取
        if ext in text_extensions or ext not in {'.pdf', '.doc', '.xls', '.ppt', '.zip', '.rar', '.7z', '.tar', '.gz', '.exe', '.dll', '.so', '.dylib'}:
            return self._try_read_text_file(file_path)

        return None

    def _try_read_text_file(self, file_path: str) -> str | None:
        """尝试以文本方式读取文件。

        Args:
            file_path: 文件路径

        Returns:
            文本内容，失败返回 None
        """
        try:
            # 尝试多种编码
            for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                        if content:
                            return content
                except (UnicodeDecodeError, UnicodeError):
                    continue
                except Exception as e:
                    log(f'qq | read text file error ({encoding}): {e}')
                    return None
            return None
        except Exception as e:
            log(f'qq | try read text file error: {e}')
            return None

    def _read_office_file(self, file_path: str, ext: str) -> str | None:
        """读取 Office 文件（docx, xlsx）。

        Args:
            file_path: 文件路径
            ext: 文件扩展名

        Returns:
            文件文本内容，失败返回 None
        """
        try:
            if ext == '.docx':
                try:
                    from docx import Document
                    doc = Document(file_path)
                    text_parts = [para.text for para in doc.paragraphs if para.text.strip()]
                    return '\n'.join(text_parts)
                except ImportError:
                    log('qq | python-docx not installed, cannot read .docx files')
                    return None
                except Exception as e:
                    log(f'qq | read docx error: {e}')
                    return None
            elif ext == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    text_lines = []
                    for sheet_name in wb.sheetnames:
                        text_lines.append(f'=== Sheet: {sheet_name} ===')
                        for row in wb[sheet_name].iter_rows(values_only=True):
                            # 过滤空行
                            if any(cell is not None for cell in row):
                                text_lines.append(' | '.join(str(cell) if cell is not None else '' for cell in row))
                    return '\n'.join(text_lines)
                except ImportError:
                    log('qq | openpyxl not installed, cannot read .xlsx files')
                    return None
                except Exception as e:
                    log(f'qq | read xlsx error: {e}')
                    return None
            return None
        except Exception as e:
            log(f'qq | read office file error: {e}')
            return None

    def _format_attachments(self, attachments: list[dict]) -> str:
        """格式化附件信息为文本。

        Args:
            attachments: 附件列表，每个附件包含 filename, content_type, size, url 等字段

        Returns:
            格式化后的附件描述文本
        """
        if not attachments:
            return ''

        lines = ['[用户发送了附件]']
        for i, att in enumerate(attachments, 1):
            filename = att.get('filename', f'file_{i}')
            content_type = att.get('content_type', 'unknown')
            size = att.get('size', 0)
            url = att.get('url', '')

            # 格式化文件大小
            if size < 1024:
                size_str = f'{size} B'
            elif size < 1024 * 1024:
                size_str = f'{size / 1024:.1f} KB'
            else:
                size_str = f'{size / (1024 * 1024):.1f} MB'

            lines.append(f'  {i}. {filename} ({content_type}, {size_str})')
            if url:
                lines.append(f'     URL: {url}')

        return '\n'.join(lines)

    async def _handle_event(self, event: dict) -> None:
        """处理从 WebSocket 接收到的事件。

        Args:
            event: 事件数据
        """
        op = event.get('op')
        event_type = event.get('t')
        data = event.get('d', {})
        seq = event.get('s')

        # 更新序列号
        if seq is not None:
            self._seq = seq

        # 处理不同操作码
        if op == 0:  # Dispatch - 事件下发
            if event_type == 'READY':
                # 连接成功
                self._session_id = data.get('session_id', '')
                log(f'qq | ready, session_id: {self._session_id}')

            elif event_type in ('AT_MESSAGE_CREATE', 'GROUP_AT_MESSAGE_CREATE'):
                # 群聊 @消息
                await self._handle_group_message(data)

            elif event_type in ('C2C_MESSAGE_CREATE', 'DIRECT_MESSAGE_CREATE'):
                # 私聊消息
                await self._handle_private_message(data)

            elif event_type == 'RESUMED':
                log('qq | resumed session')

        elif op == 10:  # Hello - 连接成功，需要发送 Identify
            log(f'qq | hello received, interval: {data.get("heartbeat_interval")}')
            await self._identify()

        elif op == 11:  # Heartbeat ACK
            pass  # 心跳响应，无需处理

        elif op == 7:  # Reconnect
            log('qq | reconnect requested')
            # 简单处理：断开连接触发重连
            if self._ws:
                await self._ws.close()

        elif op == 9:  # Invalid Session
            log('qq | invalid session, re-identifying')
            await self._identify()

    async def _handle_group_message(self, data: dict) -> None:
        """处理群聊消息事件。

        Args:
            data: 事件数据
        """
        try:
            msg_id = data.get('id', '')
            content = data.get('content', '').strip()
            group_openid = data.get('group_openid', '')
            author = data.get('author', {})
            author_id = author.get('member_openid', '')
            attachments = data.get('attachments', [])

            chat_id = f'group:{group_openid}'

            log(f'qq | group message from {author_id} in {group_openid}: {content}')

            # 忽略机器人自己的消息
            if author.get('bot', False):
                return

            # 处理附件（用户发送的文件）
            if attachments:
                try:
                    # 下载并处理每个附件
                    for attachment in attachments:
                        await self._handle_attachment(attachment, chat_id, msg_id, author_id)
                    # 附件处理完成后直接返回，不再放入空文本消息
                    return
                except Exception as e:
                    log(f'qq | attachment processing error: {e}')
                    import traceback
                    log(f'qq | traceback: {traceback.format_exc()}')
                    # 即使附件处理失败，也继续处理文本内容

            # 放入消息队列（仅在有文本内容或附件处理失败时）
            if content or not attachments:
                queue = self._get_queue(chat_id)
                await queue.put({
                    'chat_id': chat_id,
                    'msg_id': msg_id,
                    'content': content,
                    'author_id': author_id,
                    'group_openid': group_openid,
                    'type': 'group'
                })
                self._notify_message_received()

        except Exception as e:
            log(f'qq | handle group message error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    async def _handle_private_message(self, data: dict) -> None:
        """处理私聊消息事件。

        Args:
            data: 事件数据
        """
        try:
            msg_id = data.get('id', '')
            content = data.get('content', '').strip()
            author = data.get('author', {})
            user_openid = data.get('author', {}).get('user_openid', '')
            attachments = data.get('attachments', [])

            # 私聊使用 guild_id 标识会话
            guild_id = data.get('guild_id', '')

            chat_id = f'private:{user_openid}'

            log(f'qq | private message from {user_openid}: {content}')

            # 忽略机器人自己的消息
            if author.get('bot', False):
                return

            # 处理附件（用户发送的文件）
            if attachments:
                try:
                    # 下载并处理每个附件
                    for attachment in attachments:
                        await self._handle_attachment(attachment, chat_id, msg_id, user_openid)
                    # 附件处理完成后直接返回，不再放入空文本消息
                    return
                except Exception as e:
                    log(f'qq | attachment processing error: {e}')
                    import traceback
                    log(f'qq | traceback: {traceback.format_exc()}')
                    # 即使附件处理失败，也继续处理文本内容

            # 放入消息队列（仅在有文本内容或附件处理失败时）
            if content or not attachments:
                queue = self._get_queue(chat_id)
                await queue.put({
                    'chat_id': chat_id,
                    'msg_id': msg_id,
                    'content': content,
                    'author_id': user_openid,
                    'guild_id': guild_id,
                    'type': 'private'
                })
                self._notify_message_received()

        except Exception as e:
            log(f'qq | handle private message error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    async def _send_qq_message(self, chat_id: str, text: str, msg_id: str = '') -> None:
        """通过 REST API 发送 QQ 消息。

        Args:
            chat_id: 聊天 ID (group:xxx 或 private:xxx)
            text: 消息文本
            msg_id: 被动回复的消息 ID
        """
        try:
            log(f'qq | _send_qq_message starting, chat_id={chat_id}, text_len={len(text)}')

            # 确保有有效的 token
            await self._get_access_token()

            headers = self._get_auth_header()

            # 解析 chat_id
            if chat_id.startswith('group:'):
                # 群消息
                group_openid = chat_id.split(':', 1)[1]
                url = f'{self._get_api_base()}/v2/groups/{group_openid}/messages'
                payload = {
                    'msg_type': 0,  # 文本消息
                    'content': text,
                }
            elif chat_id.startswith('private:'):
                # 私聊消息
                user_openid = chat_id.split(':', 1)[1]
                url = f'{self._get_api_base()}/v2/users/{user_openid}/messages'
                payload = {
                    'msg_type': 0,  # 文本消息
                    'content': text,
                }
            else:
                log(f'qq | invalid chat_id: {chat_id}')
                return

            # 如果有 msg_id，添加为被动回复
            if msg_id:
                payload['msg_id'] = msg_id

            # 分割长消息（QQ 限制 4096 字符）
            max_len = 2000
            if len(text) > max_len:
                parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
            else:
                parts = [text]

            # 发送消息
            async with aiohttp.ClientSession() as session:
                for i, part in enumerate(parts):
                    async with session.post(url, headers=headers, json=payload) as resp:
                        if resp.status != 200:
                            error_text = await resp.text()
                            log(f'qq | send message part {i+1}/{len(parts)} failed: {resp.status} {error_text}')
                        else:
                            result = await resp.json()
                            log(f'qq | message sent, msg_id: {result.get("id")}')

                    # 避免超频，添加延迟
                    if len(parts) > 1:
                        await asyncio.sleep(1)

            log(f'qq | send completed')

        except Exception as e:
            log(f'qq | send message error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    def _send_qq_message_sync(self, chat_id: str, text: str, msg_id: str = '') -> None:
        """同步发送 QQ 消息（从主线程调用）。

        Args:
            chat_id: 聊天 ID
            text: 消息文本
            msg_id: 被动回复的消息 ID
        """
        if self._event_loop and self._event_loop.is_running():
            # 在事件循环中运行异步函数
            asyncio.run_coroutine_threadsafe(
                self._send_qq_message(chat_id, text, msg_id),
                self._event_loop
            ).result(timeout=30)
        else:
            # 直接运行（启动阶段）
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(self._send_qq_message(chat_id, text, msg_id))
            finally:
                loop.close()

    def _send_welcome_and_system(self, chat_id: str) -> None:
        """发送欢迎消息到用户和 system 提示给 Agent。

        Args:
            chat_id: 聊天 ID
        """
        if chat_id in self._welcome_sent_chats:
            return

        self._welcome_sent_chats.add(chat_id)

        # 发送欢迎消息
        welcome_msg = (
            "Successfully connected to QQ\n"
            "Welcome back! This is Momoka~\n"
            "For more help, type /help"
        )
        self._send_qq_message_sync(chat_id, welcome_msg)
        log(f'qq | welcome message sent to {chat_id}')

        # 发送 system 消息给 Agent
        if self._agent is not None:
            system_msg = "<You are communicating with user via QQ Bot>"
            self._agent._model._ctx.history.append({
                'role': 'system',
                'content': system_msg
            })
            self._agent._model._ctx._meta.append({})
            log(f'qq | system message added to agent history')

    def _get_message(self, chat_id: str) -> dict | None:
        """从指定聊天的消息队列中获取消息。

        Args:
            chat_id: 聊天 ID

        Returns:
            消息数据，队列为空时返回 None
        """
        queue = self._message_queues.get(chat_id)
        if queue and not queue.empty():
            try:
                return queue.get_nowait()
            except asyncio.QueueEmpty:
                pass
        return None

    def _run_websocket_loop(self) -> None:
        """在独立线程中运行 WebSocket 循环。"""
        self._event_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self._event_loop)

        async def connect_and_listen():
            """连接网关并监听事件。"""
            while self._running:
                try:
                    # 获取网关 URL
                    ws_url = await self._get_gateway_url()

                    # 建立 WebSocket 连接
                    async with websockets.connect(ws_url) as ws:
                        self._ws = ws
                        log('qq | websocket connected')

                        # 监听事件
                        while self._running:
                            try:
                                message = await asyncio.wait_for(ws.recv(), timeout=30)
                                event = json.loads(message)
                                await self._handle_event(event)
                            except asyncio.TimeoutError:
                                # 发送心跳
                                await self._send_heartbeat()
                            except websockets.ConnectionClosed:
                                log('qq | websocket connection closed')
                                break

                except Exception as e:
                    log(f'qq | websocket error: {e}')
                    if self._running:
                        log('qq | reconnecting in 5 seconds...')
                        await asyncio.sleep(5)

        try:
            self._event_loop.run_until_complete(connect_and_listen())
        except Exception as e:
            log(f'qq | websocket loop error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    def run(self) -> None:
        """启动 QQ Bot 会话循环。"""
        if self._agent is None:
            raise RuntimeError("Agent not set. Call set_agent() before run().")

        new_log()
        log('qq_bot_start')

        self._start_time = time.time()
        self.session.reset()
        self._running = True

        # 在后台线程启动 WebSocket
        self._ws_thread = threading.Thread(target=self._run_websocket_loop, daemon=True)
        self._ws_thread.start()

        # 等待 WebSocket 连接就绪
        while not self._session_id:
            time.sleep(0.1)

        log('qq | bot ready, waiting for messages...')
        print("Connection successful.")

        # 主线程处理消息循环
        self._run_main_loop()

    def _run_main_loop(self) -> None:
        """主消息循环: 等待用户输入,调用 Agent,返回结果。"""
        import asyncio as aio

        # 创建斜杠命令处理器
        from user.qq_bot.interactions import QQInteractionAdapter
        adapter = QQInteractionAdapter(self)
        slash_handler = adapter.create_callbacks()

        # 模型选择状态
        waiting_for_model_choice = False
        pending_models: list[str] = []

        # 清空确认状态
        waiting_for_clear_confirmation = False

        # 最后一个 msg_id（用于被动回复）
        last_msg_id: str = ''

        # 初始化消息事件（用于事件驱动）
        self._message_event = asyncio.Event()

        while self._running:
            # 等待新消息事件（替代轮询）
            try:
                # 设置超时防止永久阻塞
                self._message_event.wait(timeout=1.0)
                self._message_event.clear()
            except Exception:
                pass

            # 轮询所有队列（事件驱动 + 轮询兜底）
            active_chat_id = None
            for chat_id, queue in self._message_queues.items():
                if not queue.empty():
                    active_chat_id = chat_id
                    log(f'qq | found message in {chat_id}')
                    break

            if active_chat_id is None:
                continue

            queue = self._message_queues[active_chat_id]

            try:
                msg_data = queue.get_nowait()
            except aio.QueueEmpty:
                continue

            # 设置当前活跃的聊天 ID
            self._active_chat_id = active_chat_id
            user_message = msg_data['content']
            last_msg_id = msg_data.get('msg_id', '')

            log(f'qq | processing message: {user_message[:50]}...')

            if self._agent is None:
                continue

            # 首条消息:发送欢迎消息和 system 提示
            self._send_welcome_and_system(active_chat_id)

            # 如果正在等待模型选择,处理用户的数字回复
            if waiting_for_model_choice:
                waiting_for_model_choice = False
                if self._handle_model_choice(user_message, pending_models):
                    continue

            # 如果正在等待清空确认,处理用户的回复
            if waiting_for_clear_confirmation:
                waiting_for_clear_confirmation = False
                if self._handle_clear_confirmation(user_message):
                    continue

            # 检查是否为斜杠命令
            if user_message.strip().startswith('/'):
                if user_message.strip() == '/model':
                    models = self._fetch_models_for_model()
                    if models:
                        pending_models = models
                        waiting_for_model_choice = True
                        cfg = get_config()
                        current = cfg.get('model', '')
                        model_list = '\n'.join([
                            f"  {i+1}. {m}{' [current]' if m == current else ''}"
                            for i, m in enumerate(models)
                        ])
                        self._send_qq_message_sync(
                            active_chat_id,
                            f"Available models:\n{model_list}\n\n"
                            f"Please reply with the model number (1-{len(models)}) or leave blank to cancel.",
                            last_msg_id
                        )
                        continue

                # 处理斜杠命令
                from user.commands import SlashCommandCallbacks
                handled = self._handle_slash_command(
                    user_message, slash_handler, active_chat_id, last_msg_id
                )
                if handled:
                    continue

            # 修复历史
            repaired = self._agent.repair_history()
            if repaired:
                log(f'qq | repair_history: filled in {repaired} orphaned tool_calls')

            # 调用 Agent
            try:
                result = self._agent.send(user_message)

                self.session.update(result)

                if result.is_finish:
                    self._agent.finish_task()
                    self.on_task_finish()

            except Exception as e:
                log(f'qq | agent error: {e}')
                self._send_qq_message_sync(
                    active_chat_id,
                    f"Error: {str(e)}",
                    last_msg_id
                )

    def _handle_slash_command(
        self,
        message: str,
        handler,
        chat_id: str,
        msg_id: str
    ) -> bool:
        """处理斜杠命令。

        Args:
            message: 用户消息
            handler: 斜杠命令处理器
            chat_id: 聊天 ID
            msg_id: 消息 ID

        Returns:
            是否已处理
        """
        from user.commands import SlashCommandCallbacks

        # 简化处理：直接调用 handler
        # 实际应该通过 SlashCommandCallbacks 处理
        msg_stripped = message.strip()

        if msg_stripped == '/end':
            self._handle_end_command()
            return True

        if msg_stripped == '/help':
            self._send_qq_message_sync(
                chat_id,
                "Available commands:\n"
                "  /end - End session\n"
                "  /usage - Show token usage\n"
                "  /config - Show config\n"
                "  /model - Change model\n"
                "  /clear - Clear context\n"
                "  /help - Show this help",
                msg_id
            )
            return True

        if msg_stripped == '/usage':
            elapsed = time.time() - self._start_time
            stats = (
                f"Session stats:\n"
                f"  Input tokens: {self.session.input_tokens}\n"
                f"  Output tokens: {self.session.output_tokens}\n"
                f"  Rounds: {self.session.round_count}\n"
                f"  Elapsed: {elapsed:.0f}s"
            )
            self._send_qq_message_sync(chat_id, stats, msg_id)
            return True

        return False

    def _handle_model_choice(self, user_message: str, models: list[str]) -> bool:
        """处理模型选择。

        Args:
            user_message: 用户输入
            models: 可用模型列表

        Returns:
            是否已处理
        """
        try:
            choice = user_message.strip()
            if not choice:
                self._send_qq_message_sync(self._active_chat_id, 'Model selection cancelled.')
                return True

            idx = int(choice)
            if 1 <= idx <= len(models):
                from config import get_config, CONFIG_FILE
                cfg = get_config()
                cfg['model'] = models[idx - 1]
                with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(cfg, f, ensure_ascii=False, indent=2)
                self._send_qq_message_sync(
                    self._active_chat_id,
                    f'Model changed to: {models[idx - 1]}'
                )
                return True
            else:
                self._send_qq_message_sync(
                    self._active_chat_id,
                    f'Invalid selection. Please enter 1-{len(models)}.'
                )
                return True
        except ValueError:
            return False

    def _handle_clear_confirmation(self, user_message: str) -> bool:
        """处理清空确认。"""
        if user_message.strip().lower() in ('y', 'yes', '确认', '是'):
            if self._agent is not None:
                self._agent.clear_context()
                self.session.reset()
                self._send_qq_message_sync(self._active_chat_id, 'Context cleared.')
            return True
        return False

    def _handle_end_command(self) -> None:
        """处理结束命令。"""
        self._running = False
        if self._ws:
            asyncio.run_coroutine_threadsafe(
                self._ws.close(),
                self._event_loop
            )
        self.on_session_end(
            self.session.input_tokens,
            self.session.output_tokens,
            self.session.round_count,
            time.time() - self._start_time
        )

    def _fetch_models_for_model(self) -> list[str]:
        """获取可用模型列表。"""
        try:
            from model.model import fetch_available_models
            return fetch_available_models()
        except Exception as e:
            log(f'qq | fetch models error: {e}')
            return []

    def get_input(self) -> str:
        """等待并获取用户输入(阻塞式)。"""
        import asyncio as aio
        while True:
            for queue in self._message_queues.values():
                if not queue.empty():
                    msg = queue.get_nowait()
                    return msg['content']
            time.sleep(0.1)

    def send_output(self, message: str, role: str = 'BOT') -> None:
        """向用户输出消息。"""
        log(f'qq ({role}) | {message}')

        if self._active_chat_id:
            # QQ 添加角色标签，与飞书/Discord 保持一致
            if role == 'BOT':
                self._send_qq_message_sync(self._active_chat_id, message)
            else:
                self._send_qq_message_sync(self._active_chat_id, f"[{role}]\n{message}")

    def send_error(self, message: str) -> None:
        """输出错误消息。"""
        log(f'qq ERROR | {message}')
        if self._active_chat_id:
            self._send_qq_message_sync(self._active_chat_id, f"Error: {message}")

    def send_file(self, file_path: str, caption: str = '') -> None:
        """向用户发送一个文件（支持 QQ 富媒体消息）。

        Args:
            file_path: 文件的绝对路径
            caption: 可选的伴随消息
        """
        import os
        import base64

        log(f'qq | send_file: {file_path} | caption: {caption}')

        if not os.path.exists(file_path):
            log(f'qq | send_file error: file not found: {file_path}')
            self.send_error(f'File not found: {file_path}')
            return

        if not os.path.isfile(file_path):
            log(f'qq | send_file error: not a file: {file_path}')
            self.send_error(f'Not a file: {file_path}')
            return

        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)

        # 检查文件大小（QQ 限制，通常 25MB）
        max_size = 25 * 1024 * 1024  # 25MB
        if file_size > max_size:
            log(f'qq | send_file error: file too large: {file_size} bytes')
            self.send_error(f'File too large (max 25MB): {file_name}')
            return

        # 判断是否为群聊或私聊
        if not self._active_chat_id:
            log('qq | send_file error: no active chat')
            self.send_error('No active chat to send file to')
            return

        # 尝试使用富媒体消息发送
        try:
            self._send_file_via_media_sync(self._active_chat_id, file_path, caption)
        except Exception as e:
            log(f'qq | send_file via media failed: {e}')
            # 降级为文本消息
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')
            msg = f"[File: {file_name}]\n{caption}" if caption else f"[File: {file_name}]"
            self._send_qq_message_sync(self._active_chat_id, msg)

    def _send_file_via_media_sync(self, chat_id: str, file_path: str, caption: str = '') -> None:
        """通过 QQ 富媒体接口发送文件。

        Args:
            chat_id: 聊天 ID (group:xxx 或 private:xxx)
            file_path: 文件路径
            caption: 伴随消息
        """
        import base64

        # 读取文件并编码为 base64
        with open(file_path, 'rb') as f:
            file_data = f.read()

        file_base64 = base64.b64encode(file_data).decode('utf-8')
        file_name = os.path.basename(file_path)

        # 第一步：上传文件
        upload_result = self._upload_file_media_sync(chat_id, file_base64, file_path, file_name)
        if not upload_result:
            raise Exception('Failed to upload file to QQ')

        file_info = upload_result.get('file_info')
        if not file_info:
            raise Exception('No file_info returned from upload')

        # 第二步：发送富媒体消息
        self._send_media_message_sync(chat_id, file_info, caption or file_name)

    def _upload_file_media_sync(self, chat_id: str, file_base64: str, file_path: str, file_name: str = '') -> dict | None:
        """上传文件到 QQ 富媒体服务器。

        Args:
            chat_id: 聊天 ID
            file_base64: base64 编码的文件数据
            file_path: 文件路径（用于判断文件类型）
            file_name: 文件名（用于文件类型上传时显示）

        Returns:
            上传结果字典，包含 file_info 字段
        """
        import asyncio

        # 判断文件类型
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ('.png', '.jpg', '.jpeg'):
            file_type = 1  # 图片
        elif ext in ('.mp4',):
            file_type = 2  # 视频
        elif ext in ('.silk', '.wav', '.mp3', '.flac'):
            file_type = 3  # 语音
        else:
            file_type = 4  # 文件

        # 构建上传 URL
        if chat_id.startswith('group:'):
            group_openid = chat_id.split(':', 1)[1]
            url = f'{self._get_api_base()}/v2/groups/{group_openid}/files'
        elif chat_id.startswith('private:'):
            user_openid = chat_id.split(':', 1)[1]
            url = f'{self._get_api_base()}/v2/users/{user_openid}/files'
        else:
            log(f'qq | upload_file error: invalid chat_id: {chat_id}')
            return None

        # 准备请求
        headers = self._get_auth_header()
        headers['Content-Type'] = 'application/json'

        payload = {
            'file_type': file_type,
            'srv_send_msg': False,  # 仅上传，不直接发送
            'file_data': file_base64,
        }

        # 对于文件类型，添加文件名信息
        if file_type == 4 and file_name:
            payload['file_name'] = file_name

        log(f'qq | uploading file: type={file_type}, name={file_name or "unknown"}, size: {len(file_base64)} bytes (base64)')

        try:
            if self._event_loop and self._event_loop.is_running():
                # 在事件循环中运行
                future = asyncio.run_coroutine_threadsafe(
                    self._upload_file_request(url, headers, payload),
                    self._event_loop
                )
                return future.result(timeout=60)
            else:
                # 直接运行
                loop = asyncio.new_event_loop()
                try:
                    return loop.run_until_complete(
                        self._upload_file_request(url, headers, payload)
                    )
                finally:
                    loop.close()
        except Exception as e:
            log(f'qq | upload_file error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')
            return None

    async def _upload_file_request(self, url: str, headers: dict, payload: dict) -> dict | None:
        """执行上传文件的 HTTP 请求。

        Args:
            url: 上传 URL
            headers: 请求头
            payload: 请求体

        Returns:
            响应字典
        """
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=payload, timeout=aiohttp.ClientTimeout(total=60)) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    log(f'qq | upload_file failed: {resp.status} {error_text}')
                    return None
                else:
                    result = await resp.json()
                    log(f'qq | file uploaded, file_info: {result.get("file_info", "")[:50]}...')
                    return result

    def _send_media_message_sync(self, chat_id: str, file_info: str, caption: str = '') -> None:
        """发送富媒体消息（使用已上传的 file_info）。

        Args:
            chat_id: 聊天 ID
            file_info: 文件信息字符串
            caption: 伴随消息
        """
        import asyncio

        # 构建发送 URL
        if chat_id.startswith('group:'):
            group_openid = chat_id.split(':', 1)[1]
            url = f'{self._get_api_base()}/v2/groups/{group_openid}/messages'
            payload = {
                'content': caption,  # 群聊必填
                'msg_type': 7,  # Media 富媒体消息
                'media': {
                    'file_info': file_info,
                },
            }
        elif chat_id.startswith('private:'):
            user_openid = chat_id.split(':', 1)[1]
            url = f'{self._get_api_base()}/v2/users/{user_openid}/messages'
            payload = {
                'msg_type': 7,  # Media 富媒体消息
                'media': {
                    'file_info': file_info,
                },
                'content': caption,  # 单聊选填，但建议提供
            }
        else:
            log(f'qq | send_media_message error: invalid chat_id: {chat_id}')
            return

        # 添加被动回复的 msg_id（如果有）
        # 注意：这里我们使用最新的消息 ID 作为被动回复
        # 但在工具调用场景下，我们可能没有直接的 msg_id，所以可以省略

        log(f'qq | sending media message (msg_type=7), caption: {caption}')

        try:
            if self._event_loop and self._event_loop.is_running():
                # 在事件循环中运行
                asyncio.run_coroutine_threadsafe(
                    self._send_media_request(url, payload),
                    self._event_loop
                ).result(timeout=30)
            else:
                # 直接运行
                loop = asyncio.new_event_loop()
                try:
                    loop.run_until_complete(self._send_media_request(url, payload))
                finally:
                    loop.close()
        except Exception as e:
            log(f'qq | send_media_message error: {e}')
            import traceback
            log(f'qq | traceback: {traceback.format_exc()}')

    async def _send_media_request(self, url: str, payload: dict) -> None:
        """执行发送富媒体消息的 HTTP 请求。

        Args:
            url: 发送 URL
            payload: 请求体
        """
        headers = self._get_auth_header()
        headers['Content-Type'] = 'application/json'

        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=payload) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    log(f'qq | send_media_message failed: {resp.status} {error_text}')
                else:
                    result = await resp.json()
                    log(f'qq | media message sent, msg_id: {result.get("id")}')
