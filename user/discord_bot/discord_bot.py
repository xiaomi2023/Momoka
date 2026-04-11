"""
user/discord_bot/ —— Discord Bot 用户交互实现。

基于 discord.py 库实现 Discord 机器人接口。
支持斜杠命令、消息格式化等功能。
"""

from __future__ import annotations

import asyncio
import threading
import time
from typing import TYPE_CHECKING

import discord
from discord.ext import commands

from logger import log, new_log
from user.user import BaseUser
from config import get_config

if TYPE_CHECKING:
    from host.momoka import Momoka


class DiscordBotUser(BaseUser):
    """基于 Discord Bot 的用户交互实现。

    支持:
    - 私聊
    - 群聊 @提及 回复
    - 命令交互 (/start, /end, /usage 等)
    - 消息格式化(角色标签、长消息分割)
    """
    
    interface_type = 'discord'

    def __init__(self, token: str, allowed_users: list[int] | None = None, proxy: str | None = None):
        """
        Args:
            token: Discord Bot Token
            allowed_users: 允许使用的用户 ID 列表, None 表示允许所有用户
            proxy: 代理地址 (如 http://127.0.0.1:7890)
        """
        super().__init__()
        self._agent: Momoka | None = None
        self._start_time = 0.0
        self._token = token
        self._allowed_users = allowed_users or []
        self._proxy = proxy
        self._bot: commands.Bot | None = None
        self._user_queues: dict[int, asyncio.Queue] = {}  # channel_id -> queue
        self._event_loop: asyncio.AbstractEventLoop | None = None
        self._bot_thread: threading.Thread | None = None
        
        # 当前活跃的消息频道 ID
        self._active_channel_id: int | None = None
        
        # 是否已发送过欢迎消息(每个频道独立追踪)
        self._welcome_sent_channels: set[int] = set()

    def set_agent(self, agent: Momoka) -> None:
        """设置关联的 Agent 实例。"""
        self._agent = agent

    def _is_user_allowed(self, user_id: int) -> bool:
        """检查用户是否被允许使用。"""
        if not self._allowed_users:
            return True
        return user_id in self._allowed_users

    def _get_queue(self, channel_id: int) -> asyncio.Queue:
        """获取或创建对应 channel_id 的消息队列。"""
        if channel_id not in self._user_queues:
            self._user_queues[channel_id] = asyncio.Queue()
        return self._user_queues[channel_id]

    async def _send_discord_message(self, channel: discord.abc.Messageable, text: str) -> None:
        """发送消息到 Discord。

        Args:
            channel: Discord 消息目标(频道/私聊)
            text: 消息文本
        """
        try:
            log(f'discord | _send_discord_message starting, channel={channel}, text_len={len(text)}')
            # Discord 单条消息 2000 字符限制
            max_len = 1900
            if len(text) > max_len:
                # 分割长消息
                parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
                for i, part in enumerate(parts):
                    if i > 0:
                        part = f"{part}"
                    # 使用代码块发送长文本,避免格式化问题
                    if len(part) > max_len:
                        part = part[:max_len] + "..."
                    log(f'discord | sending part {i+1}/{len(parts)}')
                    await channel.send(part)
            else:
                log(f'discord | sending message (single part)')
                await channel.send(text)
            log(f'discord | send completed')
        except Exception as e:
            log(f'discord | send error: {e}')
            try:
                # 失败时使用纯文本
                max_len = 1900
                await channel.send(text[:max_len] if len(text) > max_len else text)
            except Exception as e2:
                log(f'discord | fallback send error: {e2}')

    def _send_discord_message_by_api(self, channel_id: int, text: str) -> None:
        """通过 Discord REST API 发送消息(不依赖 Bot 事件循环)。

        Args:
            channel_id: 频道 ID
            text: 消息文本
        """
        import aiohttp
        import asyncio as aio

        try:
            log(f'discord | _send_discord_message_by_api starting, channel_id={channel_id}, text_len={len(text)}')
            
            # Discord 单条消息 2000 字符限制
            max_len = 1900
            parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
            
            async def send_part(part_text: str, part_num: int, total: int):
                proxy = self._proxy if self._proxy else None
                async with aiohttp.ClientSession() as session:
                    async with session.post(
                        f'https://discord.com/api/v10/channels/{channel_id}/messages',
                        headers={
                            'Authorization': f'Bot {self._token}',
                            'Content-Type': 'application/json'
                        },
                        json={'content': part_text},
                        proxy=proxy
                    ) as resp:
                        log(f'discord | API response part {part_num}/{total}: {resp.status}')
                        if resp.status != 200:
                            error_text = await resp.text()
                            log(f'discord | API response body: {error_text}')

            # 同步调用异步函数
            loop = aio.new_event_loop()
            for i, part in enumerate(parts):
                loop.run_until_complete(send_part(part, i + 1, len(parts)))
            loop.close()
            log(f'discord | send completed')
        except Exception as e:
            log(f'discord | send error: {e}')
            import traceback
            log(f'discord | traceback: {traceback.format_exc()}')

    def _format_message_with_role(self, message: str, role: str) -> str:
        """根据角色格式化消息。

        Args:
            message: 消息内容
            role: 角色名称

        Returns:
            格式化后的消息
        """
        if role == 'BOT':
            return message
        return f"[{role}]\n{message}"

    def _send_welcome_and_system(self, channel_id: int) -> None:
        """发送欢迎消息到用户和 system 提示给 Agent。

        Args:
            channel_id: 频道 ID
        """
        if channel_id in self._welcome_sent_channels:
            return
        
        self._welcome_sent_channels.add(channel_id)
        
        # 发送欢迎消息到用户
        welcome_msg = (
            "Successfully connected to Discord\n"
            "Welcome back! This is Momoka~\n"
            "Developed by Mikoris | For more help, type /help"
        )
        self._send_discord_message_by_api(channel_id, welcome_msg)
        log(f'discord | welcome message sent to channel {channel_id}')
        
        # 发送 system 消息给 Agent(添加到对话历史,不触发响应)
        if self._agent is not None:
            system_msg = "<You are communicating with user via Discord>"
            self._agent._model._ctx.history.append({
                'role': 'system',
                'content': system_msg
            })
            self._agent._model._ctx._meta.append({})
            log(f'discord | system message added to agent history')

    async def _on_ready(self):
        """Bot 启动完成回调。"""
        log(f'discord | logged in as {self._bot.user}')
        log(f'discord | Bot ID: {self._bot.user.id}')

    async def _on_message(self, message: discord.Message):
        """消息接收回调。"""
        # 忽略自己
        if message.author == self._bot.user:
            return

        # 检查权限
        if not self._is_user_allowed(message.author.id):
            await message.channel.send("Error: You do not have permission to use this bot.")
            return

        content = message.content.strip()
        log(f'discord | received message from {message.author}: {content}')

        # 处理附件（文件）
        if message.attachments:
            await self._handle_attachments(message)
            return

        # 普通消息 - 放入队列
        if content:
            queue = self._get_queue(message.channel.id)
            await queue.put(content)
            log(f'discord | message queued to channel {message.channel.id}')

    async def _handle_attachments(self, message: discord.Message):
        """处理消息附件（文件）。

        Args:
            message: Discord 消息对象
        """
        try:
            channel_id = message.channel.id
            log(f'discord | handling {len(message.attachments)} attachment(s) from channel {channel_id}')

            for attachment in message.attachments:
                file_name = attachment.filename
                log(f'discord | processing attachment: {file_name}')

                # 下载文件
                file_path = await self._download_discord_attachment(attachment)
                if file_path:
                    # 读取文件内容
                    file_content = self._read_file_content(file_path, file_name)
                    if file_content:
                        # 将文件内容作为消息放入队列
                        message_text = f"[User uploaded file: {file_name}]\n\n{file_content}"
                        queue = self._get_queue(channel_id)
                        await queue.put(message_text)
                        log(f'discord | file content queued: {file_name}')
                    else:
                        queue = self._get_queue(channel_id)
                        await queue.put(f"[User uploaded file: {file_name}]\n[File downloaded to: {file_path} but could not be read as text]")
                else:
                    log(f'discord | failed to download attachment: {file_name}')
                    await message.channel.send(f"Failed to receive file: {file_name}")
        except Exception as e:
            log(f'discord | handle attachments error: {e}')
            import traceback
            log(f'discord | traceback: {traceback.format_exc()}')

    async def _download_discord_attachment(self, attachment: discord.Attachment) -> str | None:
        """下载 Discord 附件。

        Args:
            attachment: Discord 附件对象

        Returns:
            下载后的本地文件路径，失败返回 None
        """
        try:
            import os
            import tempfile
            import aiohttp

            # 使用工作目录下的 temp 文件夹
            work_dir = get_config().get('work_dir', tempfile.gettempdir())
            temp_dir = os.path.join(work_dir, 'temp')
            os.makedirs(temp_dir, exist_ok=True)

            # 生成唯一文件名避免冲突
            import uuid
            file_name = attachment.filename
            unique_name = f"{uuid.uuid4().hex[:8]}_{file_name}"
            file_path = os.path.join(temp_dir, unique_name)

            # 下载文件
            proxy = self._proxy if self._proxy else None
            async with aiohttp.ClientSession() as session:
                async with session.get(attachment.url, proxy=proxy) as resp:
                    if resp.status != 200:
                        log(f'discord | download attachment failed: {resp.status}')
                        return None
                    with open(file_path, 'wb') as f:
                        f.write(await resp.read())

            log(f'discord | attachment downloaded to: {file_path}')
            return file_path

        except Exception as e:
            log(f'discord | download attachment error: {e}')
            import traceback
            log(f'discord | traceback: {traceback.format_exc()}')
            return None

    def _read_file_content(self, file_path: str, file_name: str) -> str | None:
        """读取文件内容为文本。

        Args:
            file_path: 文件路径
            file_name: 文件名

        Returns:
            文件文本内容，如果不是文本文件返回 None
        """
        import os

        # 检查文件大小（限制 10MB）
        max_size = 10 * 1024 * 1024  # 10MB
        file_size = os.path.getsize(file_path)
        if file_size > max_size:
            return f"[File too large: {file_size / 1024 / 1024:.2f}MB, max 10MB]"

        # 根据扩展名判断文件类型
        ext = os.path.splitext(file_name)[1].lower()

        # 文本文件扩展名
        text_extensions = {
            '.txt', '.md', '.py', '.js', '.ts', '.html', '.css', '.json',
            '.xml', '.yaml', '.yml', '.csv', '.log', '.ini', '.conf',
            '.sh', '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd',
            '.c', '.cpp', '.h', '.hpp', '.java', '.go', '.rs', '.swift',
            '.kt', '.scala', '.rb', '.php', '.pl', '.lua', '.r', '.m',
            '.sql', '.dockerfile', '.makefile', '.cmake', '.gradle',
            '.vue', '.jsx', '.tsx', '.svelte', '.less', '.scss', '.sass',
        }

        # Office 文档扩展名（需要特殊处理）
        office_extensions = {
            '.docx', '.xlsx', '.pptx',
        }

        if ext in office_extensions:
            return self._read_office_file(file_path, ext)

        # 尝试作为文本文件读取
        if ext in text_extensions or ext not in {'.pdf', '.doc', '.xls', '.ppt', '.zip', '.rar', '.7z', '.tar', '.gz', '.exe', '.dll', '.so', '.dylib'}:
            return self._try_read_text_file(file_path)

        return None

    def _try_read_text_file(self, file_path: str) -> str | None:
        """尝试以文本方式读取文件。

        Args:
            file_path: 文件路径

        Returns:
            文件内容，如果读取失败返回 None
        """
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1', 'cp1252']

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                return content
            except UnicodeDecodeError:
                continue
            except Exception as e:
                log(f'discord | read file error with {encoding}: {e}')
                continue

        return None

    def _read_office_file(self, file_path: str, ext: str) -> str | None:
        """读取 Office 文档内容。

        Args:
            file_path: 文件路径
            ext: 文件扩展名

        Returns:
            文档文本内容
        """
        try:
            if ext == '.docx':
                from docx import Document
                doc = Document(file_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return '\n'.join(paragraphs)
            elif ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                lines = []
                for sheet in wb.worksheets:
                    lines.append(f'--- Sheet: {sheet.title} ---')
                    for row in sheet.iter_rows(values_only=True):
                        row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                        if row_text.strip():
                            lines.append(row_text)
                return '\n'.join(lines)
            elif ext == '.pptx':
                # python-pptx 需要单独安装，这里简单返回提示
                return "[PPTX file - content extraction not supported]"
        except Exception as e:
            log(f'discord | read office file error: {e}')
            return f"[Failed to read {ext} file: {e}]"

        return None

    def run(self) -> None:
        """启动 Discord Bot 会话循环。"""
        if self._agent is None:
            raise RuntimeError("Agent not set. Call set_agent() before run().")

        new_log()
        log('discord_bot_start')

        self._start_time = time.time()
        self.session.reset()

        def run_bot():
            """在独立线程中运行 Discord Bot。"""
            self._event_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(self._event_loop)

            # 设置代理（monkey patch discord.http 的 HTTPClient）
            if self._proxy:
                import functools
                log(f'discord | using proxy: {self._proxy}')

                # 保存原始的 __init__ 方法
                original_http_init = discord.http.HTTPClient.__init__
                proxy_url = self._proxy  # 捕获当前值

                def patched_http_init(http_self, *args, **kwargs):
                    # 强制注入 proxy 参数
                    kwargs['proxy'] = proxy_url
                    original_http_init(http_self, *args, **kwargs)

                # 应用 patch
                discord.http.HTTPClient.__init__ = patched_http_init

            # 设置 Intents
            intents = discord.Intents.default()
            intents.message_content = True
            # intents.members = True  # 特权意图，需要在 Developer Portal 启用
            intents.members = False  # 禁用特权意图

            # 创建 Bot
            self._bot = commands.Bot(command_prefix='!', intents=intents)

            # 注册事件（使用 bot 内部方法注册）
            @self._bot.event
            async def on_ready():
                await self._on_ready()

            @self._bot.event
            async def on_message(message: discord.Message):
                await self._on_message(message)

            log('discord | bot starting')
            self._bot.run(self._token)

        # 在后台线程启动 Bot
        self._bot_thread = threading.Thread(target=run_bot, daemon=True)
        self._bot_thread.start()

        # 主线程处理消息循环
        self._run_main_loop()

    def _run_main_loop(self) -> None:
        """主消息循环: 等待用户输入,调用 Agent,返回结果。"""
        import asyncio as aio

        # 创建斜杠命令处理器
        from user.discord_bot.interactions import DiscordInteractionAdapter
        adapter = DiscordInteractionAdapter(self)
        slash_handler = self._create_slash_handler(adapter)

        # 模型选择状态
        waiting_for_model_choice = False
        pending_models: list[str] = []

        # 清空确认状态
        waiting_for_clear_confirmation = False

        # 等待 Bot 启动
        while self._bot is None or not self._bot.is_ready():
            time.sleep(0.1)

        log('discord | bot ready, waiting for messages...')

        while True:
            # 轮询所有队列
            active_channel_id = None
            for channel_id, queue in self._user_queues.items():
                if not queue.empty():
                    active_channel_id = channel_id
                    log(f'discord | found message in channel {channel_id}')
                    break

            if active_channel_id is None:
                time.sleep(0.1)
                continue

            queue = self._user_queues[active_channel_id]

            try:
                user_message = queue.get_nowait()
                log(f'discord | processing message: {user_message[:50]}...')
            except aio.QueueEmpty:
                continue

            # 设置当前活跃的频道 ID
            self._active_channel_id = active_channel_id

            if self._agent is None:
                continue

            # 首条消息:发送欢迎消息和 system 提示
            self._send_welcome_and_system(active_channel_id)

            # 如果正在等待模型选择,处理用户的数字回复
            if waiting_for_model_choice:
                waiting_for_model_choice = False
                if self._handle_model_choice(user_message, pending_models):
                    continue
                # 如果选择无效,继续作为普通消息处理

            # 如果正在等待清空确认,处理用户的回复
            if waiting_for_clear_confirmation:
                waiting_for_clear_confirmation = False
                if self._handle_clear_confirmation(user_message):
                    continue

            # 检查是否为斜杠命令
            if user_message.strip().startswith('/'):
                # 如果是 /model 命令,需要特殊处理
                if user_message.strip() == '/model':
                    models = self._fetch_models_for_model()
                    if models:
                        pending_models = models
                        waiting_for_model_choice = True
                        # 发送模型列表
                        cfg = get_config()
                        current = cfg.get('model', '')
                        model_list = '\n'.join([
                            f"  {i+1}. {m}{' [current]' if m == current else ''}"
                            for i, m in enumerate(models)
                        ])
                        self._send_discord_message_by_api(
                            active_channel_id,
                            f"Available models:\n{model_list}\n\n"
                            f"Please reply with the model number (1-{len(models)}) or leave blank to cancel."
                        )
                        continue
                    else:
                        self._send_discord_message_by_api(active_channel_id, 'No models available or failed to fetch.')
                        continue

                handled, skill_name = slash_handler.handle(user_message)
                if handled:
                    # /end 命令特殊处理
                    if skill_name == '__end__':
                        self._handle_end_command()
                        break
                    # /init 命令
                    if skill_name == '__init__':
                        continue
                    # /clear 命令（等待用户确认）
                    if skill_name == '__clear_ask__':
                        waiting_for_clear_confirmation = True
                        continue
                    # Skill 加载
                    if skill_name is not None:
                        log(f'discord | skill trigger: {skill_name}')
                        load_result = self._agent.load_skill(skill_name)
                        if load_result.success:
                            self._send_discord_message_by_api(active_channel_id, f'Load Skill: {skill_name}')
                            log(f'discord (skill inject): {skill_name}')
                        else:
                            self._send_discord_message_by_api(active_channel_id, f'Non-existent command or skill: {skill_name}')
                        continue
                    # 其他命令已处理
                    continue

            # 修复历史
            repaired = self._agent.repair_history()
            if repaired:
                log(f'discord | repair_history: filled in {repaired} orphaned tool_calls')

            # 调用 Agent
            try:
                result = self._agent.send(
                    user_message,
                    file_contents=self.session.file_contents
                )

                self.session.update(result)

                if result.is_finish:
                    self._agent.finish_task()
                    self.on_task_finish()

            except Exception as e:
                log(f'discord | agent error: {e}')
                if self._bot is not None and self._event_loop is not None:
                    channel = self._bot.get_channel(active_channel_id)
                    if channel:
                        try:
                            future = asyncio.run_coroutine_threadsafe(
                                self._send_discord_message(channel, f"Error: {str(e)}"),
                                self._event_loop
                            )
                            future.result(timeout=10)
                        except Exception as send_err:
                            log(f'discord | send error message failed: {send_err}')

    def get_input(self) -> str:
        """等待并获取用户输入(阻塞式)。"""
        import asyncio as aio
        while True:
            for queue in self._user_queues.values():
                if not queue.empty():
                    return queue.get_nowait()
            time.sleep(0.1)

    def send_output(self, message: str, role: str = 'BOT') -> None:
        """向用户输出消息。"""
        log(f'discord ({role}) | {message}')

        # 如果有活跃的频道 ID,发送消息到 Discord
        if self._active_channel_id:
            # 格式化消息
            formatted_msg = self._format_message_with_role(message, role)
            # 发送消息
            self._send_discord_message_by_api(self._active_channel_id, formatted_msg)

    def send_error(self, message: str) -> None:
        """输出错误消息。"""
        log(f'discord ERROR | {message}')

    def send_file(self, file_path: str, caption: str = '') -> None:
        """向用户发送一个文件（Discord 支持文件附件）。

        Args:
            file_path: 文件的绝对路径
            caption: 可选的伴随消息
        """
        import os

        log(f'discord | send_file: {file_path} | caption: {caption}')

        # 验证文件是否存在
        if not os.path.exists(file_path):
            log(f'discord | send_file error: file not found: {file_path}')
            self.send_error(f'File not found: {file_path}')
            return

        # 如果有活跃的频道 ID,发送文件到 Discord
        if self._active_channel_id:
            self._send_discord_file_by_api(self._active_channel_id, file_path, caption)

    def _send_discord_file_by_api(self, channel_id: int, file_path: str, caption: str = '') -> None:
        """通过 Discord REST API 发送文件附件。

        Args:
            channel_id: 频道 ID
            file_path: 文件路径
            caption: 可选的伴随消息
        """
        import aiohttp
        import asyncio as aio
        import os

        try:
            file_name = os.path.basename(file_path)
            log(f'discord | _send_discord_file_by_api starting, channel_id={channel_id}, file={file_name}')

            # Discord 附件上传使用 multipart/form-data
            proxy = self._proxy if self._proxy else None

            async def send_file_async():
                async with aiohttp.ClientSession() as session:
                    # 构建 multipart 数据
                    data = aiohttp.FormData()
                    if caption:
                        data.add_field('content', caption)
                    
                    with open(file_path, 'rb') as f:
                        data.add_field(
                            'files[0]',
                            f,
                            filename=file_name,
                            content_type='application/octet-stream'
                        )

                        async with session.post(
                            f'https://discord.com/api/v10/channels/{channel_id}/messages',
                            headers={'Authorization': f'Bot {self._token}'},
                            data=data,
                            proxy=proxy
                        ) as resp:
                            log(f'discord | file upload response: {resp.status}')
                            if resp.status not in (200, 201):
                                error_text = await resp.text()
                                log(f'discord | file upload error: {error_text}')

            # 同步调用异步函数
            loop = aio.new_event_loop()
            loop.run_until_complete(send_file_async())
            loop.close()
            log(f'discord | file sent: {file_name}')

        except Exception as e:
            log(f'discord | send file error: {e}')
            import traceback
            log(f'discord | traceback: {traceback.format_exc()}')
            self.send_error(f'Failed to send file: {e}')

    def on_task_finish(self) -> None:
        """任务完成回调。"""
        log('discord | task finished')

    def on_session_end(self, input_tokens: int, output_tokens: int,
                       round_count: int, elapsed: float) -> None:
        """会话结束回调。"""
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)
        time_str = f'{mins}min {secs}s' if mins else f'{secs}s'
        log(f'discord | session end ({time_str} | Input: {input_tokens} tokens | Output: {output_tokens} tokens | {round_count}R)')

    def _create_slash_handler(self, adapter) -> 'SlashCommandHandler':
        """创建斜杠命令处理器。"""
        from user.commands import SlashCommandHandler
        return SlashCommandHandler(adapter.create_callbacks())

    def _handle_end_command(self) -> None:
        """处理 /end 命令。"""
        elapsed = time.time() - self._start_time
        self.on_session_end(
            self.session.input_tokens,
            self.session.output_tokens,
            self.session.round_count,
            elapsed
        )

    def _fetch_models_for_model(self) -> list[str]:
        """为 /model 命令获取模型列表。"""
        try:
            from openai import OpenAI
            cfg = get_config()
            client = OpenAI(api_key=cfg['api_key'], base_url=cfg['base_url'])
            models = client.models.list()
            return sorted(m.id for m in models.data)
        except Exception as e:
            log(f'discord | failed to fetch models: {e}')
            return []

    def _handle_model_choice(self, user_message: str, models: list[str]) -> bool:
        """处理用户的模型选择输入。

        Args:
            user_message: 用户回复的消息
            models: 模型列表

        Returns:
            True 表示已处理，False 表示输入无效
        """
        user_input = user_message.strip()

        # 空输入视为取消
        if not user_input:
            self._send_discord_message_by_api(self._active_channel_id, 'Cancelled.')
            return True

        # 尝试解析数字
        try:
            choice = int(user_input)
        except ValueError:
            # 不是数字，视为无效，返回 False 让主循环继续作为普通消息处理
            return False

        # 验证范围
        if choice < 1 or choice > len(models):
            self._send_discord_message_by_api(
                self._active_channel_id,
                f'Invalid selection. Please enter a number between 1 and {len(models)}.'
            )
            return True

        selected_model = models[choice - 1]
        cfg = get_config()
        current = cfg.get('model', '')

        # 如果选择的模型和当前相同，无需更新
        if selected_model == current:
            self._send_discord_message_by_api(
                self._active_channel_id,
                f'Model unchanged: {selected_model}'
            )
            return True

        # 更新配置
        try:
            import json
            from config import CONFIG_FILE
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config_data = json.load(f)

            config_data['model'] = selected_model

            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=2)

            self._send_discord_message_by_api(
                self._active_channel_id,
                f'Model updated: {current!r} → {selected_model!r}'
            )
            log(f'discord | model changed: {current} → {selected_model}')
            return True

        except Exception as e:
            self._send_discord_message_by_api(
                self._active_channel_id,
                f'Failed to save model: {e}'
            )
            return True

    def _handle_clear_confirmation(self, user_message: str) -> bool:
        """处理用户对清空操作的确认回复。

        Args:
            user_message: 用户回复的消息

        Returns:
            True 表示已处理
        """
        user_input = user_message.strip().lower()

        # 检查是否确认
        if user_input in ('y', 'yes'):
            try:
                # 清空模型层的对话历史
                if self._agent:
                    self._agent.clear_context()

                    # 重置会话状态（token 统计等）
                    self.session.reset()

                    # 触发回调
                    self.on_clear_context()

                self._send_discord_message_by_api(self._active_channel_id, 'Context cleared.')
                return True

            except Exception as e:
                self._send_discord_message_by_api(self._active_channel_id, f'Failed to clear context: {e}')
                return True
        else:
            # 用户取消
            self._send_discord_message_by_api(self._active_channel_id, 'Cancelled.')
            return True
